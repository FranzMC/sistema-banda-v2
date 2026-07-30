from rest_framework import viewsets, views, status, permissions
from rest_framework.response import Response
from rest_framework.decorators import action
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import date
from decimal import Decimal
import uuid

from rest_framework_simplejwt.tokens import RefreshToken
from django.db import transaction

from .. import services
from ..models import (
    Usuario, Musico, Evento, Asistencia, Descuento, Pago,
    RendimientoMusico, ConfiguracionSistema, Adelanto, PlanillaLiquidacion,
    ContratoMusico, DetalleMontoDiario, JefeSeccion, Deuda, AbonoDeuda,
    CampanaCanaston, ResultadoCanaston
)
from ..permissions import EsAdministrativo, EsJefeSeccion, EsPresidente
from ..serializers import (
    UsuarioSerializer, UsuarioCreateSerializer,
    MusicoSerializer, MusicoListSerializer, EventoSerializer,
    AsistenciaSerializer, DescuentoSerializer, PagoSerializer,
    ConfiguracionSistemaSerializer, AdelantoSerializer,
    PlanillaLiquidacionSerializer, PlanillaLiquidacionDetalleSerializer,
    CampanaCanastonSerializer, ResultadoCanastonSerializer
)
from services.rendimiento_calculator import RendimientoCalculator
from ..canaston_service import calcular_elegibilidad_canaston

class UsuarioViewSet(viewsets.ModelViewSet):
    def get_queryset(self):
        queryset = Usuario.objects.all().order_by('rol', 'last_name', 'first_name')
        musico_id = self.request.query_params.get('musico_id', None)
        if musico_id:
            queryset = queryset.filter(perfil_musico__id=musico_id)
        return queryset
    permission_classes = [permissions.IsAuthenticated, EsPresidente]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return UsuarioCreateSerializer
        return UsuarioSerializer

    @action(detail=False, methods=['get'])
    def roles_modules(self, request):
        """Devuelve los módulos permitidos para cada rol definidos en la base de datos"""
        from ..models import RolModulo
        roles_dict = {}
        # Obtener todos los mapeos de rol a módulo
        asignaciones = RolModulo.objects.select_related('modulo').filter(modulo__activo=True)
        
        for asignacion in asignaciones:
            if asignacion.rol not in roles_dict:
                roles_dict[asignacion.rol] = []
            roles_dict[asignacion.rol].append(asignacion.modulo.nombre)
            
        # Asegurarse de que al menos existan listas vacías para roles sin módulos
        for rol, _ in Usuario.ROLES:
            if rol not in roles_dict:
                roles_dict[rol] = []
                
        return Response(roles_dict)

    @action(detail=True, methods=['post'])
    def reset_pin(self, request, pk=None):
        """Restablece el PIN del usuario usando los 4 primeros dígitos de su CI"""
        usuario = self.get_object()
        ci = None
        
        # Intentar obtener CI del perfil de músico si existe
        if hasattr(usuario, 'perfil_musico') and usuario.perfil_musico:
            ci = usuario.perfil_musico.documento_identidad
            
        if not ci:
            return Response({'error': 'El usuario no tiene un CI asociado o no tiene perfil de músico.'}, status=status.HTTP_400_BAD_REQUEST)
            
        # Generar contraseña usando los primeros 4 dígitos del CI
        ci_digits = ''.join(ch for ch in str(ci) if ch.isdigit())
        if len(ci_digits) >= 4:
            base = ci_digits[:4]
        else:
            base = ci_digits.ljust(4, '0')
            
        usuario.set_password(base)
        usuario.pin_actual = base
        usuario.save()
        
        return Response({'success': True, 'message': 'PIN restablecido exitosamente al valor por defecto.'})

    @action(detail=False, methods=['post'])
    def update_order(self, request):
        order_data = request.data.get('order', [])
        for item in order_data:
            Musico.objects.filter(id=item['id']).update(orden=item['order'])
        return Response({'success': True})

    @action(detail=False, methods=['get'])
    def generar_plantilla_excel(self, request):
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return Response({'error': 'openpyxl no está instalado'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Descuentos"
        
        # ---------------------------------------------------------
        # CONFIGURACIÓN DE PÁGINA PARA PDF
        # ---------------------------------------------------------
        # Orientación Horizontal (Landscape) para tener el doble de espacio horizontal
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        
        # Márgenes exactos de 1 cm (0.3937 pulgadas) en todos los lados
        ws.page_margins.left = 0.3937
        ws.page_margins.right = 0.3937
        ws.page_margins.top = 0.3937
        ws.page_margins.bottom = 0.3937
        ws.page_margins.header = 0.3937
        ws.page_margins.footer = 0.3937
        
        # Centrar la tabla horizontalmente en la hoja
        ws.print_options.horizontalCentered = True
        
        # Auto-escalado: Encoge todo lo necesario para que el ancho entre en 1 sola página
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # ---------------------------------------------------------
        # ESTILOS PROFESIONALES
        # ---------------------------------------------------------
        # Fuentes principales
        main_title_font = Font(name='Arial', size=16, bold=True, color="000000")
        
        # Estilos de Cabecera (Fondo Azul Oscuro Institucional, Letra Blanca)
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        
        # Estilos de Sección (Fondo Gris Claro, Letra Negra)
        section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        section_font = Font(name='Arial', size=10, bold=True, color="000000")
        
        # Estilos de Datos
        data_font = Font(name='Arial', size=9, color="000000")
        total_font = Font(name='Arial', size=9, bold=True, color="000000")
        
        # Bordes Negros Delgados bien definidos (Obligatorio para que no falle el lector PDF)
        thin_border = Border(
            left=Side(style='thin', color='000000'), 
            right=Side(style='thin', color='000000'), 
            top=Side(style='thin', color='000000'), 
            bottom=Side(style='thin', color='000000')
        )
        
        # ---------------------------------------------------------
        # TÍTULO PRINCIPAL
        # ---------------------------------------------------------
        ws.merge_cells('A1:G2')
        ws['A1'] = "Descuento Pdf"
        ws['A1'].font = main_title_font
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # ---------------------------------------------------------
        # CABECERAS Y ROTACIÓN
        # ---------------------------------------------------------
        headers = ['N°', 'NOMBRES Y APELLIDOS', 'ATRASO', 'FALTA', 'UNIFORME', 'BEBIDAS', 'TOTAL']
        
        # Aumentamos el alto de la fila para que el texto hacia arriba tenga espacio
        ws.row_dimensions[4].height = 80 
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            
            # Si son columnas de descuento o total (de la 3 en adelante), rotamos 90 grados y permitimos salto de línea
            if col_num > 2:
                cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # ---------------------------------------------------------
        # ANCHOS DE COLUMNA
        # ---------------------------------------------------------
        ws.column_dimensions['A'].width = 4   # N°
        ws.column_dimensions['B'].width = 30  # Nombres (Espacio holgado para nombres largos)
        
        # Las columnas de descuentos y total ahora son ajustadas, ancho 6 permite dos líneas de texto rotado
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 6
            
        # ---------------------------------------------------------
        # DATOS DE LOS MÚSICOS
        # ---------------------------------------------------------
        musicos = Musico.objects.filter(activo=True).order_by('orden', 'apellidos')
        secciones = ['TROMPETA', 'CLARINETE', 'SAXOFON', 'BARITONO', 'TROMBON', 'TUBA', 'BOMBO', 'TAMBOR', 'PLATILLOS', 'PERCUSION', 'OTRO']
        
        current_row = 5
        for seccion in secciones:
            musicos_seccion = [m for m in musicos if m.instrumento == seccion]
            if not musicos_seccion: continue
            
            # Fila separadora de sección
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            cell = ws.cell(row=current_row, column=1, value=seccion)
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            
            # Borde para la celda combinada
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = thin_border
                
            current_row += 1
            
            for idx, musico in enumerate(musicos_seccion, 1):
                # N°
                c1 = ws.cell(row=current_row, column=1, value=idx)
                c1.border = thin_border
                c1.font = data_font
                c1.alignment = Alignment(horizontal="center", vertical="center")
                
                # Nombres
                c2 = ws.cell(row=current_row, column=2, value=musico.nombre_completo)
                c2.border = thin_border
                c2.font = data_font
                c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                
                # Espacios vacíos para ingresar números
                for col in range(3, 7):
                    cv = ws.cell(row=current_row, column=col, value='')
                    cv.border = thin_border
                    cv.font = data_font
                    cv.alignment = Alignment(horizontal="center", vertical="center")
                
                # TOTAL (Fórmula en negrita)
                ctot = ws.cell(row=current_row, column=7, value=f"=SUM(C{current_row}:F{current_row})")
                ctot.font = total_font
                ctot.border = thin_border
                ctot.alignment = Alignment(horizontal="center", vertical="center")
                
                current_row += 1
                
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Plantilla_Multas_Adelantos.xlsx'
        wb.save(response)
        return response
