from rest_framework import viewsets, views, status, permissions
from rest_framework.response import Response
from rest_framework.decorators import action
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Count, Exists, OuterRef, Subquery, BooleanField, CharField, IntegerField
from django.utils import timezone
from datetime import date
from decimal import Decimal
import uuid

from rest_framework_simplejwt.tokens import RefreshToken
from django.db import transaction

from .. import services
from ..models import (
    Usuario, Musico, Evento, Asistencia, Descuento, Pago,
    RendimientoMusico, ConfiguracionSistema, Adelanto, PlanillaLiquidacion,
    ContratoMusico, DetalleMontoDiario, JefeSeccion, Deuda, AbonoDeuda,
    CampanaCanaston, ResultadoCanaston
)
from ..permissions import EsAdministrativo, EsJefeSeccion, EsPresidente
from ..serializers import (
    UsuarioSerializer, UsuarioCreateSerializer,
    MusicoSerializer, MusicoListSerializer, EventoSerializer,
    AsistenciaSerializer, DescuentoSerializer, PagoSerializer,
    ConfiguracionSistemaSerializer, AdelantoSerializer,
    PlanillaLiquidacionSerializer, PlanillaLiquidacionDetalleSerializer,
    CampanaCanastonSerializer, ResultadoCanastonSerializer
)
from services.rendimiento_calculator import RendimientoCalculator
from ..canaston_service import calcular_elegibilidad_canaston

class MusicoViewSet(viewsets.ModelViewSet):
    queryset = Musico.objects.filter(activo=True).order_by('orden', 'apellidos', 'nombres')
    serializer_class = MusicoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        """
        Selecciona el serializer basado en la acción y los parámetros de la consulta.
        - `response=light`: Usa el serializer ligero (para la app móvil).
        - Por defecto: Usa el serializer completo (para el frontend web).
        """
        if self.request.query_params.get('response') == 'light':
            return MusicoListSerializer

        return MusicoSerializer

    def check_permissions(self, request):
        super().check_permissions(request)
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            if request.user.rol in ['MUSICO']:
                self.permission_denied(request, message="No tienes permisos para modificar músicos.")

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()

        # Filtrar por sección según el rol del usuario
        if user.rol == 'JEFE_SECCION' and self.request.query_params.get('todas_secciones') != 'true':
            # El jefe de sección SOLO ve los músicos de su sección
            instrumentos_permitidos = user.get_instrumentos_encargados()
            if instrumentos_permitidos:
                queryset = queryset.filter(instrumento__in=instrumentos_permitidos)
            else:
                return queryset.none()
        # MUSICO: solo se ve a sí mismo (opcional, lo dejamos ver todos por ahora)
        # Dirección, Presidente, etc: ven todo

        # Luego, aplicar los filtros de búsqueda y de instrumento si están presentes.
        search = self.request.query_params.get('search', None)
        instrumento = self.request.query_params.get('instrumento', None)
        
        if search:
            queryset = queryset.filter(
                Q(nombres__icontains=search) | Q(apellidos__icontains=search) | Q(documento_identidad__icontains=search)
            )
        
        # El filtro por instrumento solo aplica para roles directivos (no para JEFE_SECCION que ya está filtrado)
        if user.rol in ['PRESIDENTE', 'DIRECTOR', 'SUBDIRECTOR'] or user.is_superuser:
            if instrumento:
                queryset = queryset.filter(instrumento=instrumento)
                
        fecha = self.request.query_params.get('disponible_en_fecha')
        excluir_evento_id = self.request.query_params.get('excluir_evento_id')
        if fecha:
            # Anotar si está ocupado en algún evento en esa fecha
            from datetime import datetime
            try:
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
                eventos_mismo_dia = Evento.objects.filter(
                    fecha_hora_cita__date=fecha_obj,
                    convocados=OuterRef('pk')
                )
                if excluir_evento_id:
                    eventos_mismo_dia = eventos_mismo_dia.exclude(id=excluir_evento_id)
                    
                queryset = queryset.annotate(
                    ocupado_en_fecha=Exists(eventos_mismo_dia),
                    evento_ocupado_titulo=Subquery(eventos_mismo_dia.values('titulo')[:1], output_field=CharField()),
                    evento_ocupado_id=Subquery(eventos_mismo_dia.values('id')[:1], output_field=IntegerField())
                )
            except ValueError:
                pass
            
        return queryset

    def perform_create(self, serializer):
        """
        Crea un músico y su usuario asociado usando la capa de servicios.
        Los jefes de sección solo pueden crear músicos de su sección.
        """
        user = self.request.user
        
        # Mapeo de instrumentos a categorías para percusión
        percusion_instruments = ['BOMBO', 'PLATILLOS', 'TAMBOR', 'PERCUSION']
        
        # Validar que jefe de sección solo cree músicos de su sección
        if user.rol == 'JEFE_SECCION':
            instrumento = serializer.validated_data.get('instrumento')
            instrumentos_permitidos = user.get_instrumentos_encargados()
            
            if instrumento not in instrumentos_permitidos:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied(f'Solo puede crear músicos de sus secciones: {", ".join(instrumentos_permitidos)}')
        
        validated_data = serializer.validated_data
        nombres = validated_data.get('nombres', '')
        apellidos = validated_data.get('apellidos', '')
        ci = validated_data.get('documento_identidad', '')
        
        user_created = services.create_user_for_musico(
            nombres=nombres,
            apellidos=apellidos,
            ci=ci
        )
        
        serializer.save(usuario=user_created)

    def perform_update(self, serializer):
        """
        Actualiza un músico.
        Los jefes de sección solo pueden editar músicos de su sección.
        """
        user = self.request.user
        
        # Mapeo de instrumentos a categorías para percusión
        percusion_instruments = ['BOMBO', 'PLATILLOS', 'TAMBOR', 'PERCUSION']
        
        # Validar que jefe de sección solo edite músicos de su sección
        if user.rol == 'JEFE_SECCION':
            musico = self.get_object()
            instrumentos_permitidos = user.get_instrumentos_encargados()
            
            if musico.instrumento not in instrumentos_permitidos:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied(f'Solo puede editar músicos de sus secciones: {", ".join(instrumentos_permitidos)}')
            
            # Validar que no cambie el instrumento a otra sección
            instrumento = serializer.validated_data.get('instrumento')
            if instrumento and instrumento not in instrumentos_permitidos:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied(f'No puede cambiar el instrumento a otra sección fuera de: {", ".join(instrumentos_permitidos)}')
        
        serializer.save()

    def perform_destroy(self, instance):
        """
        Elimina un músico.
        Los jefes de sección solo pueden eliminar músicos de su sección.
        """
        user = self.request.user
        
        # Mapeo de instrumentos a categorías para percusión
        percusion_instruments = ['BOMBO', 'PLATILLOS', 'TAMBOR', 'PERCUSION']
        
        # Validar que jefe de sección solo elimine músicos de su sección
        if user.rol == 'JEFE_SECCION':
            instrumentos_permitidos = user.get_instrumentos_encargados()
            
            if instance.instrumento not in instrumentos_permitidos:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied(f'Solo puede eliminar músicos de sus secciones: {", ".join(instrumentos_permitidos)}')
        
        instance.delete()

    @action(detail=False, methods=['post'])
    def update_order(self, request):
        order_data = request.data.get('order', [])
        for item in order_data:
            Musico.objects.filter(id=item['id']).update(orden=item['order'])
        return Response({'success': True})

    @action(detail=False, methods=['get'])
    def generar_plantilla_excel(self, request):
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
            
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return Response({'error': 'openpyxl no está instalado'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Descuentos"
        
        # ---------------------------------------------------------
        # CONFIGURACIÓN DE PÁGINA PARA PDF
        # ---------------------------------------------------------
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.3937
        ws.page_margins.right = 0.3937
        ws.page_margins.top = 0.3937
        ws.page_margins.bottom = 0.3937
        ws.page_margins.header = 0.3937
        ws.page_margins.footer = 0.3937
        ws.print_options.horizontalCentered = True
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        main_title_font = Font(name='Arial', size=16, bold=True, color="000000")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        section_font = Font(name='Arial', size=10, bold=True, color="000000")
        data_font = Font(name='Arial', size=9, color="000000")
        total_font = Font(name='Arial', size=9, bold=True, color="000000")
        thin_border = Border(
            left=Side(style='thin', color='000000'), 
            right=Side(style='thin', color='000000'), 
            top=Side(style='thin', color='000000'), 
            bottom=Side(style='thin', color='000000')
        )
        
        ws.merge_cells('A1:G2')
        ws['A1'] = "Descuento Pdf"
        ws['A1'].font = main_title_font
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ['N°', 'NOMBRES Y APELLIDOS', 'ATRASO', 'FALTA', 'UNIFORME', 'BEBIDAS', 'TOTAL']
        ws.row_dimensions[4].height = 80 
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            if col_num > 2:
                cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 30
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 6
        
        musicos = Musico.objects.filter(activo=True).order_by('orden', 'apellidos')
        secciones = ['TROMPETA', 'CLARINETE', 'SAXOFON', 'BARITONO', 'TROMBON', 'TUBA', 'BOMBO', 'TAMBOR', 'PLATILLOS', 'PERCUSION', 'OTRO']
        current_row = 5
        for seccion in secciones:
            musicos_seccion = [m for m in musicos if m.instrumento == seccion]
            if not musicos_seccion:
                continue
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            cell = ws.cell(row=current_row, column=1, value=seccion)
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1
            for idx, musico in enumerate(musicos_seccion, 1):
                c1 = ws.cell(row=current_row, column=1, value=idx)
                c1.border = thin_border
                c1.font = data_font
                c1.alignment = Alignment(horizontal="center", vertical="center")
                c2 = ws.cell(row=current_row, column=2, value=musico.nombre_completo)
                c2.border = thin_border
                c2.font = data_font
                c2.alignment = Alignment(horizontal="left", vertical="center")
                for col in range(3, 8):
                    cell = ws.cell(row=current_row, column=col, value="")
                    cell.border = thin_border
                current_row += 1
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plantilla_descuentos.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def reporte(self, request):
        formato = request.query_params.get('tipo_reporte', 'excel')
        columns_str = request.query_params.get('columns', '')
        columns = columns_str.split(',') if columns_str else ['nombres_apellidos', 'instrumento']

        musicos_qs = self.get_queryset()
        
        def section_weight(instr):
            s = (instr or '').upper()
            if 'TROMPETA' in s: return 1
            if 'CLARINETE' in s or 'SAXO' in s: return 2
            if 'BARITONO' in s: return 3
            if 'TROMBON' in s: return 4
            if 'TUBA' in s: return 5
            if 'BOMBO' in s or 'TAMBOR' in s or 'PLATILLO' in s or 'PERCUSION' in s: return 6
            return 7

        musicos = sorted(list(musicos_qs), key=lambda m: (section_weight(m.instrumento), m.orden if m.orden is not None else 999, m.apellidos or '', m.nombres or ''))

        from django.conf import settings
        import os
        logo_path = os.path.join(settings.BASE_DIR, 'media', 'imagenes', 'imagen_logo.jpg')

        if formato == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.drawing.image import Image as ExcelImage
            except ImportError:
                return Response({'error': 'openpyxl no está instalado'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            from django.http import HttpResponse

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Músicos"

            titulo_personalizado = request.query_params.get('titulo', '')

            # --- DISEÑO DEL ENCABEZADO EXCEL ---
            ws.merge_cells('A1:H2')
            ws['A1'] = "BANDA DE MUSICA INTERNACIONAL ESPECTACULAR MEJILLONES BOLIVIA\nEco De Los Andes"
            ws['A1'].font = Font(bold=True, size=14, color="000000")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if titulo_personalizado:
                ws.merge_cells('A3:H3')
                ws['A3'] = titulo_personalizado.upper()
                ws['A3'].font = Font(bold=True, size=12, color="000000")
                ws['A3'].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells('A4:H4')
            ws['A4'] = '"Por que la meji nunca pierde papá "'
            ws['A4'].font = Font(italic=True, size=11, color="000000")
            ws['A4'].alignment = Alignment(horizontal="center", vertical="center")

            # Espacio
            start_row = 7

            raw_headers = [col.replace('_', ' ').upper() for col in columns]
            headers = ['TALLAS (C/Ch/Z)' if h == 'TALLAS' else h for h in raw_headers]

            # Estilos de cabecera tabla
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for col_num, header_text in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_num, value=header_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

            current_row = start_row + 1
            for musico in musicos:
                row = []
                for col in columns:
                    if col == 'ci':
                        row.append(musico.documento_identidad or '')
                    elif col == 'nombres_apellidos':
                        row.append(f"{musico.nombres} {musico.apellidos}")
                    elif col == 'celular':
                        row.append(musico.telefono or '')
                    elif col == 'instrumento':
                        row.append(musico.instrumento or '')
                    elif col == 'tallas':
                        row.append(f"{musico.talla_camisa or '-'}/{musico.talla_chamarra or '-'}/{musico.numero_calzado or '-'}")
                    elif col == 'estado':
                        row.append('Activo' if musico.activo else 'Inactivo')
                    elif col == 'direccion':
                        row.append(musico.direccion or '')
                    elif col == 'fecha_nacimiento':
                        row.append(str(musico.fecha_nacimiento) if musico.fecha_nacimiento else '')
                    elif col == 'nivel':
                        row.append(musico.nivel or '')
                    else:
                        row.append('')

                for col_num, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=cell_value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", horizontal="left")
                current_row += 1

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Reporte_Musicos.xlsx'
            wb.save(response)
            return response

        elif formato == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter, portrait
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            except ImportError:
                return Response({'error': 'reportlab no está instalado'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            import io
            from django.http import HttpResponse

            buffer = io.BytesIO()
            # Hoja Vertical (Portrait)
            doc = SimpleDocTemplate(buffer, pagesize=portrait(letter),
                                    rightMargin=10, leftMargin=10, topMargin=10, bottomMargin=10)
            elements = []

            styles = getSampleStyleSheet()

            titulo_personalizado = request.query_params.get('titulo', '')

            # --- DISEÑO DEL ENCABEZADO PDF ---
            title_style = ParagraphStyle(
                'CustomTitle', parent=styles['Normal'], fontName='Helvetica-Bold',
                fontSize=16, textColor=colors.HexColor('#1e3a8a'), alignment=TA_CENTER,
                spaceAfter=5
            )
            custom_title_style = ParagraphStyle(
                'UserTitle', parent=styles['Normal'], fontName='Helvetica-Bold',
                fontSize=14, textColor=colors.HexColor('#eab308'), alignment=TA_CENTER,
                spaceAfter=10
            )
            phrase_style = ParagraphStyle(
                'PhraseStyle', parent=styles['Normal'], fontName='Helvetica-Oblique',
                fontSize=10, textColor=colors.HexColor('#475569'), alignment=TA_CENTER,
                spaceAfter=20
            )

            elements.append(Paragraph("BANDA DE MUSICA INTERNACIONAL ESPECTACULAR MEJILLONES BOLIVIA<br/>Eco De Los Andes", title_style))
            if titulo_personalizado:
                elements.append(Paragraph(titulo_personalizado.upper(), custom_title_style))

            elements.append(Paragraph('"Por que la meji nunca pierde papá "', phrase_style))
            elements.append(Spacer(1, 10))

            # Tabla de datos
            raw_headers = [col.replace('_', ' ').upper() for col in columns]
            headers = ['TALLAS (C/Ch/Z)' if h == 'TALLAS' else h for h in raw_headers]
            data = [headers]

            # Estilo más pequeño para que quepa en vertical
            small_style = ParagraphStyle('SmallStyle', parent=styles['Normal'], fontSize=8)

            for musico in musicos:
                row = []
                for col in columns:
                    if col == 'ci':
                        row.append(musico.documento_identidad or '')
                    elif col == 'nombres_apellidos':
                        row.append(f"{musico.nombres} {musico.apellidos}")
                    elif col == 'celular':
                        row.append(musico.telefono or '')
                    elif col == 'instrumento':
                        row.append(musico.instrumento or '')
                    elif col == 'tallas':
                        row.append(f"{musico.talla_camisa or '-'}/{musico.talla_chamarra or '-'}/{musico.numero_calzado or '-'}")
                    elif col == 'estado':
                        row.append('Activo' if musico.activo else 'Inactivo')
                    elif col == 'direccion':
                        row.append(musico.direccion or '')
                    elif col == 'fecha_nacimiento':
                        row.append(str(musico.fecha_nacimiento) if musico.fecha_nacimiento else '')
                    elif col == 'nivel':
                        row.append(musico.nivel or '')
                    else:
                        row.append('')

                data.append([Paragraph(str(item), small_style) if len(str(item)) > 15 else str(item) for item in row])

            # Reduce font size to fit portrait
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#eab308')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#eab308')),
            ]))
            elements.append(t)

            # Función para dibujar el pie de página
            def footer(canvas, doc):
                canvas.saveState()
                canvas.setFont('Helvetica', 8)
                canvas.drawRightString(doc.pagesize[0] - 30, 20, f"Página {doc.page}")
                canvas.restoreState()

            doc.build(elements, onFirstPage=footer, onLaterPages=footer)

            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=Reporte_Musicos.pdf'
            return response

        return Response({'error': 'Formato no soportado'}, status=status.HTTP_400_BAD_REQUEST)


class ContratoMusicoViewSet(viewsets.ModelViewSet):
    queryset = ContratoMusico.objects.all().order_by('-created_at')
    serializer_class = None
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        from ..serializers import ContratoMusicoSerializer
        return ContratoMusicoSerializer
    
    def perform_create(self, serializer):
        """Solo DIRECTOR, SUBDIRECTOR y PRESIDENTE pueden crear contratos"""
        if self.request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR', 'PRESIDENTE'] and not self.request.user.is_superuser:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Solo el directorio puede crear contratos')
        serializer.save()
    
    def perform_update(self, serializer):
        """Solo DIRECTOR, SUBDIRECTOR y PRESIDENTE pueden editar contratos"""
        if self.request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR', 'PRESIDENTE'] and not self.request.user.is_superuser:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Solo el directorio puede editar contratos')
        serializer.save()
    
    def perform_destroy(self, instance):
        """Solo DIRECTOR, SUBDIRECTOR y PRESIDENTE pueden eliminar contratos"""
        if self.request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR', 'PRESIDENTE'] and not self.request.user.is_superuser:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Solo el directorio puede eliminar contratos')
        instance.delete()
    
    @action(detail=False, methods=['post'])
    def asignar_montos_personalizados(self, request):
        """
        Permite al directorio asignar montos personalizados para un evento
        Soporta montos variables por día
        """
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
        
        evento_id = request.data.get('evento_id')
        contratos_data = request.data.get('contratos', [])
        
        if not evento_id or not contratos_data:
            return Response({'error': 'Datos incompletos'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            evento = Evento.objects.get(id=evento_id)
        except Evento.DoesNotExist:
            return Response({'error': 'Evento no existe'}, status=status.HTTP_404_NOT_FOUND)
        
        creados = 0
        actualizados = 0
        detalles_creados = 0
        
        for contrato_data in contratos_data:
            musico_id = contrato_data.get('musico_id')
            monto_diario_base = contrato_data.get('monto_diario')
            montos_diarios = contrato_data.get('montos_diarios', [])  # Lista de montos por fecha
            observaciones = contrato_data.get('observaciones', '')
            
            try:
                musico = Musico.objects.get(id=musico_id)
            except Musico.DoesNotExist:
                continue
            
            # Crear o actualizar contrato base
            contrato, created = ContratoMusico.objects.update_or_create(
                musico=musico,
                evento=evento,
                defaults={
                    'monto_diario': monto_diario_base,
                    'aprobado_por': request.user,
                    'fecha_aprobacion': timezone.now(),
                    'observaciones': observaciones
                }
            )
            
            if created:
                creados += 1
            else:
                actualizados += 1
            
            # Procesar montos diarios específicos si existen
            if montos_diarios:
                for monto_diario_data in montos_diarios:
                    fecha = monto_diario_data.get('fecha')
                    monto_especificico = monto_diario_data.get('monto')
                    motivo = monto_diario_data.get('motivo', '')
                    
                    if fecha and monto_especificico:
                        from datetime import datetime
                        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
                        
                        detalle, detalle_created = DetalleMontoDiario.objects.update_or_create(
                            contrato=contrato,
                            fecha=fecha_obj,
                            defaults={
                                'monto_asignado': monto_especificico,
                                'motivo_variacion': motivo,
                                'aprobado_por': request.user,
                                'fecha_aprobacion': timezone.now()
                            }
                        )
                        
                        if detalle_created:
                            detalles_creados += 1
                        
                        # Crear o actualizar asistencia para esa fecha específica
                        Asistencia.objects.update_or_create(
                            musico=musico,
                            evento=evento,
                            fecha_asistencia=fecha_obj,
                            defaults={
                                'monto_acordado': monto_especificico,
                                'contrato': contrato
                            }
                        )
            else:
                # Si no hay montos diarios específicos, usar el monto base para la fecha del evento
                fecha_evento = evento.fecha_hora_cita.date()
                Asistencia.objects.update_or_create(
                    musico=musico,
                    evento=evento,
                    fecha_asistencia=fecha_evento,
                    defaults={
                        'monto_acordado': monto_diario_base,
                        'contrato': contrato
                    }
                )
        
        return Response({
            'success': True,
            'creados': creados,
            'actualizados': actualizados,
            'detalles_diarios': detalles_creados,
            'evento': evento.titulo
        })

    @action(detail=False, methods=['post'])
    def asignar_montos_variables(self, request):
        """
        Asigna montos variables para músicos en diferentes fechas de un mismo evento
        """
        if request.user.rol not in ['DIRECTOR', 'SUBDIRECTOR'] and not request.user.is_superuser:
            return Response({'error': 'Sin permisos'}, status=status.HTTP_403_FORBIDDEN)
        
        evento_id = request.data.get('evento_id')
        musico_id = request.data.get('musico_id')
        montos_por_fecha = request.data.get('montos_por_fecha', [])
        
        if not evento_id or not musico_id or not montos_por_fecha:
            return Response({'error': 'Datos incompletos'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            evento = Evento.objects.get(id=evento_id)
            musico = Musico.objects.get(id=musico_id)
        except (Evento.DoesNotExist, Musico.DoesNotExist):
            return Response({'error': 'Evento o músico no existe'}, status=status.HTTP_404_NOT_FOUND)
        
        # Obtener o crear contrato base
        contrato, created = ContratoMusico.objects.get_or_create(
            musico=musico,
            evento=evento,
            defaults={
                'monto_diario': Decimal('0.00'),
                'aprobado_por': request.user,
                'fecha_aprobacion': timezone.now()
            }
        )
        
        detalles_procesados = 0
        for monto_data in montos_por_fecha:
            fecha = monto_data.get('fecha')
            monto = monto_data.get('monto')
            motivo = monto_data.get('motivo', '')
            
            if fecha and monto:
                from datetime import datetime
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
                
                # Crear detalle diario
                detalle, detalle_created = DetalleMontoDiario.objects.update_or_create(
                    contrato=contrato,
                    fecha=fecha_obj,
                    defaults={
                        'monto_asignado': monto,
                        'motivo_variacion': motivo,
                        'aprobado_por': request.user,
                        'fecha_aprobacion': timezone.now()
                    }
                )
                
                if detalle_created:
                    detalles_procesados += 1
                
                # Crear asistencia específica
                Asistencia.objects.update_or_create(
                    musico=musico,
                    evento=evento,
                    fecha_asistencia=fecha_obj,
                    defaults={
                        'monto_acordado': monto,
                        'contrato': contrato
                    }
                )
        
        return Response({
            'success': True,
            'detalles_procesados': detalles_procesados,
            'musico': musico.nombre_completo,
            'evento': evento.titulo
        })


class JefeSeccionViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar jefes de sección"""
    queryset = JefeSeccion.objects.all().order_by('-fecha_nombramiento')
    serializer_class = None  # Se definirá dinámicamente
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['seccion', 'activo']
    search_fields = ['musico__nombre_completo', 'seccion']
    ordering_fields = ['fecha_nombramiento', 'seccion', 'musico__nombre_completo']

    def get_serializer_class(self):
        from ..serializers import JefeSeccionSerializer
        return JefeSeccionSerializer

    @action(detail=False, methods=['get'])
    def por_seccion(self, request):
        """Retorna los jefes activos agrupados por sección"""
        jefes = JefeSeccion.objects.filter(activo=True).select_related('musico')
        serializer = self.get_serializer(jefes, many=True)
        return Response(serializer.data)

